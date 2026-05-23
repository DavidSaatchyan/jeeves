"""Unit tests for knowledge/catalog.py — parsers, helpers, and import orchestrator."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch
from uuid import UUID, uuid4

import pytest

from app.knowledge.catalog import (
    parse_csv,
    parse_json,
    parse_xlsx,
    parse_catalog,
    _parse_price,
    _parse_attributes,
    _normalize_headers,
    import_catalog,
)


# ── _parse_price ───────────────────────────────────────────────────────


class TestParsePrice:
    def test_dollar_decimal(self):
        assert _parse_price("29.99") == 2999

    def test_dollar_with_symbol(self):
        assert _parse_price("$49.95") == 4995

    def test_dollar_with_commas(self):
        assert _parse_price("1,299.99") == 129999

    def test_integer_as_cents(self):
        assert _parse_price("2999") == 2999

    def test_large_no_decimal_treated_as_cents(self):
        assert _parse_price("1000") == 1000

    def test_zero(self):
        assert _parse_price("0") == 0
        assert _parse_price("0.00") == 0

    def test_empty_string(self):
        assert _parse_price("") is None

    def test_whitespace_only(self):
        assert _parse_price("   ") is None

    def test_negative_value(self):
        result = _parse_price("-10.00")
        assert result == -1000

    def test_small_decimal(self):
        assert _parse_price("0.99") == 99

    def test_price_with_euro_symbol(self):
        """$ replacement only strips $, not other symbols — expect ValueError"""
        with pytest.raises(ValueError):
            _parse_price("€29.99")


# ── _parse_attributes ──────────────────────────────────────────────────


class TestParseAttributes:
    def test_key_value_pairs(self):
        result = _parse_attributes("color=red, size=M")
        assert result == {"color": "red", "size": "M"}

    def test_json_format(self):
        result = _parse_attributes('{"color": "red", "size": "M"}')
        assert result == {"color": "red", "size": "M"}

    def test_colon_separated(self):
        result = _parse_attributes("color: red, size: M")
        assert result == {"color": "red", "size": "M"}

    def test_empty_string(self):
        assert _parse_attributes("") == {}

    def test_empty_json_object(self):
        assert _parse_attributes("{}") == {}

    def test_single_attribute(self):
        result = _parse_attributes("color=red")
        assert result == {"color": "red"}

    def test_attribute_with_spaces(self):
        result = _parse_attributes("material = cotton, weight = 1.5 kg")
        assert result == {"material": "cotton", "weight": "1.5 kg"}

    def test_trailing_comma_handling(self):
        result = _parse_attributes("color=red, size=M,")
        assert result == {"color": "red", "size": "M"}


# ── parse_csv ──────────────────────────────────────────────────────────


class TestParseCsv:
    def test_valid_csv(self):
        csv_content = "name,product_id,price,stock_status,category\nWidget,WG-001,29.99,in_stock,Gadgets"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert products[0]["product_id"] == "WG-001"
        assert products[0]["price"] == 2999
        assert products[0]["stock_status"] == "in_stock"
        assert products[0]["category"] == "Gadgets"
        assert len(errors) == 0

    def test_multiple_rows(self):
        csv_content = "name,price\nA,10.99\nB,20.99\nC,30.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 3
        assert products[0]["price"] == 1099
        assert products[1]["price"] == 2099
        assert products[2]["price"] == 3099
        assert len(errors) == 0

    def test_missing_name_skipped(self):
        csv_content = "name,price\n,10.99\nWidget,20.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert len(errors) == 1
        assert "missing" in errors[0].lower()

    def test_empty_csv(self):
        products, errors = parse_csv("name,price\n")
        assert len(products) == 0
        assert len(errors) == 0

    def test_no_name_column(self):
        """No 'name' or 'title' column — returns specific error."""
        products, errors = parse_csv("product_name,price\nWidget,29.99")
        assert len(products) == 0
        assert len(errors) == 1
        assert "Could not find 'name'" in errors[0]
        assert "title" in errors[0]

    def test_unrecognised_single_column(self):
        """A single unrecognised column returns name-column error."""
        products, errors = parse_csv("justdata")
        assert len(products) == 0
        assert len(errors) == 1
        assert "Could not find 'name'" in errors[0]

    def test_csv_with_extra_columns(self):
        """Extra columns not in the parser's known set are silently ignored."""
        csv_content = "name,price,discount,notes\nX,9.99,10,Nice"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert "notes" not in products[0]  # not in parser output
        assert products[0]["name"] == "X"  # known field extracted
        assert len(errors) == 0

    def test_invalid_price(self):
        csv_content = "name,price\nWidget,not_a_number"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["price"] is None  # invalid price → None
        assert len(errors) == 1
        assert "price" in errors[0].lower()

    def test_default_values(self):
        csv_content = "name\nWidget"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["stock_status"] == "unknown"
        assert products[0]["currency"] == "USD"
        assert products[0]["attributes"] == {}

    def test_whitespace_handling(self):
        csv_content = "  name  ,  price  \n  Widget  ,  29.99  "
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert products[0]["price"] == 2999

    def test_comma_in_quoted_field(self):
        csv_content = 'name,description\nWidget,"A great product, really"'
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["description"] == "A great product, really"

    def test_duplicate_product_ids(self):
        csv_content = "name,product_id\nA,ID1\nB,ID1"
        products, errors = parse_csv(csv_content)
        assert len(products) == 2
        assert products[0]["product_id"] == "ID1"
        assert products[1]["product_id"] == "ID1"
        assert len(errors) == 0  # no dedup at parse level


class TestParseJson:
    def test_valid_array(self):
        data = '[{"name":"Widget","product_id":"WG-001","price":29.99}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert products[0]["price"] == 2999

    def test_dict_with_products_key(self):
        data = '{"products":[{"name":"A","price":10.99},{"name":"B","price":20.99}]}'
        products, errors = parse_json(data)
        assert len(products) == 2
        assert products[0]["price"] == 1099
        assert products[1]["price"] == 2099

    def test_dict_with_items_key(self):
        data = '{"items":[{"name":"X","price":15.99}]}'
        products, errors = parse_json(data)
        assert len(products) == 1

    def test_single_object_wrapped(self):
        data = '{"name":"Single","price":5.99}'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["name"] == "Single"

    def test_invalid_json(self):
        products, errors = parse_json("not json")
        assert len(products) == 0
        assert len(errors) == 1
        assert "Invalid JSON" in errors[0]

    def test_missing_name(self):
        data = '[{"price":10.99},{"name":"B","price":20.99}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["name"] == "B"
        assert len(errors) == 1

    def test_non_dict_item(self):
        data = '["string", {"name":"Real"}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert len(errors) == 1
        assert "expected object" in errors[0].lower()

    def test_empty_array(self):
        products, errors = parse_json("[]")
        assert len(products) == 0
        assert len(errors) == 0

    def test_price_as_int(self):
        data = '[{"name":"X","price":2999}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["price"] == 2999

    def test_price_as_null(self):
        data = '[{"name":"X","price":null}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["price"] is None

    def test_attributes_preserved(self):
        data = '[{"name":"X","attributes":{"color":"red","size":"M"}}]'
        products, errors = parse_json(data)
        assert products[0]["attributes"] == {"color": "red", "size": "M"}


# ── parse_xlsx ────────────────────────────────────────────────────────


class TestParseXlsx:
    def test_missing_pandas(self):
        with patch.dict("sys.modules", {"pandas": None}):
            with patch("builtins.__import__", side_effect=ImportError("no pandas")):
                products, errors = parse_xlsx(Path("dummy.xlsx"))
                assert len(products) == 0
                assert len(errors) == 1
                assert "pandas" in errors[0].lower()

    def test_file_not_found(self):
        products, errors = parse_xlsx(Path("nonexistent_file.xlsx"))
        assert len(products) == 0
        assert len(errors) == 1
        assert any("read" in e.lower() or "xlsx" in e.lower() for e in errors)

    def test_unsupported_extension(self):
        products, errors = parse_xlsx(Path("data.ods"))
        assert len(products) == 0
        assert len(errors) >= 1


# ── parse_catalog ──────────────────────────────────────────────────────


class TestParseCatalog:
    def test_csv_file(self, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\nWidget,29.99")
        products, errors = parse_catalog(p)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"

    def test_json_file(self, tmp_path):
        p = tmp_path / "products.json"
        p.write_text('[{"name":"Widget","price":29.99}]')
        products, errors = parse_catalog(p)
        assert len(products) == 1

    def test_xlsx_file_returns_error_if_not_found(self, tmp_path):
        p = tmp_path / "products.xlsx"
        # File doesn't exist, should return error path
        products, errors = parse_catalog(p)
        assert len(products) == 0
        assert len(errors) >= 1

    def test_unsupported_format(self, tmp_path):
        p = tmp_path / "data.xyz"
        p.write_text("data")
        products, errors = parse_catalog(p)
        assert len(products) == 0
        assert len(errors) == 1
        assert "unsupported" in errors[0].lower()

    def test_empty_csv_returns_no_products(self, tmp_path):
        p = tmp_path / "empty.csv"
        p.write_text("name,price\n")
        products, errors = parse_catalog(p)
        assert len(products) == 0
        assert len(errors) == 0


# ── import_catalog ────────────────────────────────────────────────────


class TestImportCatalog:
    @pytest.fixture
    def tenant_id(self) -> UUID:
        return uuid4()

    def test_successful_import(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,product_id,price\nWidget,WG-1,29.99\nGadget,GD-1,9.99")

        db = MagicMock()

        with patch("app.knowledge.catalog.rag.index_products", return_value=2) as mock_index:
            imported, errors, batch_id = import_catalog(tenant_id, p, db)

        assert imported == 2
        assert len(errors) == 0
        assert batch_id.startswith("import_")
        assert db.add.call_count == 2
        db.commit.assert_called_once()
        mock_index.assert_called_once()

    def test_import_empty_file_no_products(self, tenant_id, tmp_path):
        p = tmp_path / "empty.csv"
        p.write_text("name,price\n")

        db = MagicMock()
        imported, errors, batch_id = import_catalog(tenant_id, p, db)
        assert imported == 0
        assert "No products found" in errors[0]
        db.add.assert_not_called()

    def test_import_with_chroma_failure(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\nWidget,29.99")

        db = MagicMock()

        with patch("app.knowledge.catalog.rag.index_products", side_effect=ValueError("Chroma down")):
            imported, errors, batch_id = import_catalog(tenant_id, p, db)

        assert imported == 1  # SQL insert succeeded
        assert len(errors) == 1
        assert "Chroma" in errors[0]

    def test_import_parse_errors_collected(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\n,10.99\nWidget,not_a_price\nGadget,20.99")

        db = MagicMock()
        with patch("app.knowledge.catalog.rag.index_products", return_value=2):
            imported, errors, batch_id = import_catalog(tenant_id, p, db)

        assert imported == 2  # two valid rows
        assert len(errors) == 2  # two errors: missing name + invalid price
        assert db.add.call_count == 2

    def test_import_custom_batch_id(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\nWidget,29.99")

        db = MagicMock()
        imported, errors, batch_id = import_catalog(tenant_id, p, db, batch="my_batch_001")
        assert batch_id == "my_batch_001"

    def test_import_sets_all_product_fields(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text(
            "name,product_id,description,category,price,currency,stock_status,image_url,product_url\n"
            "Widget,WG-1,A widget,Gadgets,29.99,USD,in_stock,https://img.com/w.png,https://shop.com/w"
        )

        db = MagicMock()
        import_catalog(tenant_id, p, db)

        call_kwargs = db.add.call_args[0][0]
        assert call_kwargs.name == "Widget"
        assert call_kwargs.product_id == "WG-1"
        assert call_kwargs.description == "A widget"
        assert call_kwargs.category == "Gadgets"
        assert call_kwargs.price == 2999
        assert call_kwargs.currency == "USD"
        assert call_kwargs.stock_status == "in_stock"
        assert call_kwargs.image_url == "https://img.com/w.png"
        assert call_kwargs.product_url == "https://shop.com/w"


# ── Edge cases ──────────────────────────────────────────────────────────


class TestParsePriceEdgeCases:
    def test_very_large_value_no_decimal(self):
        assert _parse_price("100000000") == 100000000

    def test_decimal_only(self):
        assert _parse_price(".50") == 50

    def test_leading_zeros(self):
        assert _parse_price("00010.00") == 1000

    def test_dollar_sign_anywhere(self):
        assert _parse_price("10$00") == 1000

    def test_negative_cents(self):
        assert _parse_price("-100") == -10000

    def test_trailing_dot(self):
        assert _parse_price("10.") == 1000


class TestParseCsvEdgeCases:
    def test_bom_handling(self):
        csv_content = "\ufeffname,price\nWidget,29.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert len(errors) == 0

    def test_unicode_characters(self):
        csv_content = "name,description\nCafé,Crème brûlée"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Café"
        assert products[0]["description"] == "Crème brûlée"

    def test_empty_content_returns_no_header_error(self):
        products, errors = parse_csv("")
        assert len(products) == 0
        assert len(errors) >= 1

    def test_very_long_field_value(self):
        long_name = "A" * 10000
        csv_content = f"name,price\n{long_name},29.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert len(products[0]["name"]) == 10000

    def test_special_chars_in_values(self):
        csv_content = 'name,description\n"Widget <>&""",A&B'
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == 'Widget <>&"'

    def test_tab_separated_not_detected_by_default(self):
        csv_content = "name\tprice\nWidget\t29.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 0  # csv.DictReader uses comma by default, not tab

    def test_csv_with_only_header_row(self):
        csv_content = "name,price\n"
        products, errors = parse_csv(csv_content)
        assert len(products) == 0
        assert len(errors) == 0


class TestParseJsonEdgeCases:
    def test_null_description_becomes_string_none(self):
        data = '[{"name":"X","description":null}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["description"] == "None"

    def test_boolean_fields_handled(self):
        data = '[{"name":"X","active":true,"visible":false}]'
        products, errors = parse_json(data)
        assert len(products) == 1

    def test_nested_attributes_object(self):
        data = '[{"name":"X","attributes":{"dimensions":{"w":10,"h":20}}}]'
        products, errors = parse_json(data)
        assert products[0]["attributes"] == {"dimensions": {"w": 10, "h": 20}}

    def test_empty_object(self):
        products, errors = parse_json("{}")
        assert len(products) == 0
        assert len(errors) == 1  # {}, no products/items key → wraps in [data], data is empty → no name

    def test_empty_string_content(self):
        products, errors = parse_json("")
        assert len(products) == 0
        assert len(errors) == 1
        assert "Invalid JSON" in errors[0]

    def test_very_large_json_array(self):
        items = ','.join([f'{{"name":"P{i}","price":{i}.99}}' for i in range(1000)])
        data = f'[{items}]'
        products, errors = parse_json(data)
        assert len(products) == 1000
        assert len(errors) == 0

    def test_extra_fields_ignored(self):
        data = '[{"name":"X","price":10,"wholesale_price":8,"supplier":"Acme"}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert "wholesale_price" not in products[0]
        assert "supplier" not in products[0]


class TestParseXlsxEdgeCases:
    def test_xlsx_with_real_file(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["name", "product_id", "price", "stock_status"])
        ws.append(["Widget", "W001", "29.99", "in_stock"])
        ws.append(["Gadget", "G001", "49.99", "out_of_stock"])
        path = tmp_path / "products.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 2
        assert products[0]["name"] == "Widget"
        assert products[0]["price"] == 2999
        assert products[0]["stock_status"] == "in_stock"
        assert products[1]["name"] == "Gadget"
        assert products[1]["price"] == 4999
        assert len(errors) == 0

    def test_xlsx_with_missing_name_skipped(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "price"])
        ws.append(["", "10.99"])
        ws.append(["Widget", "20.99"])
        path = tmp_path / "products.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 1
        assert len(errors) == 1
        assert "missing" in errors[0].lower()

    def test_xlsx_empty_sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "price"])
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 0
        assert len(errors) >= 1  # "XLSX file is empty" — header-only sheet

    def test_xlsx_invalid_price(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "price"])
        ws.append(["Widget", "not_a_price"])
        path = tmp_path / "bad.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 1
        assert products[0]["price"] is None
        assert len(errors) == 1
        assert "price" in errors[0].lower()

    def test_xlsx_with_attributes_column(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "attributes"])
        ws.append(["Widget", "color=red,size=M"])
        ws.append(["Gadget", '{"material":"steel"}'])
        path = tmp_path / "attrs.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 2
        assert products[0]["attributes"] == {"color": "red", "size": "M"}
        assert products[1]["attributes"] == {"material": "steel"}


class TestParseCatalogEdgeCases:
    def test_parse_catalog_auto_detect_csv(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("name,price\nWidget,29.99")
        products, errors = parse_catalog(p)
        assert len(products) == 1

    def test_parse_catalog_auto_detect_json(self, tmp_path):
        p = tmp_path / "data.json"
        p.write_text('[{"name":"Widget","price":29.99}]')
        products, errors = parse_catalog(p)
        assert len(products) == 1

    def test_parse_catalog_auto_detect_xlsx(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "price"])
        ws.append(["Widget", "29.99"])
        p = tmp_path / "data.xlsx"
        wb.save(str(p))
        products, errors = parse_catalog(p)
        assert len(products) == 1
        assert products[0]["price"] == 2999


class TestHeaderAliases:
    """Header alias normalisation for _normalize_headers and all parsers."""

    # ── _normalize_headers unit ──────────────────────────────────────────

    def test_normalize_headers_exact_match(self):
        mapping = _normalize_headers(["name", "price", "category"])
        assert mapping == {"name": "name"}

    def test_normalize_headers_title_alias(self):
        mapping = _normalize_headers(["Title", "price"])
        assert mapping == {"Title": "name"}

    def test_normalize_headers_sku_alias(self):
        mapping = _normalize_headers(["name", "sku"])
        assert mapping.get("sku") == "product_id"

    def test_normalize_headers_availability_alias(self):
        mapping = _normalize_headers(["name", "availability"])
        assert mapping.get("availability") == "stock_status"

    def test_normalize_headers_desc_alias(self):
        mapping = _normalize_headers(["name", "desc"])
        assert mapping.get("desc") == "description"

    def test_normalize_headers_image_alias(self):
        mapping = _normalize_headers(["name", "image"])
        assert mapping.get("image") == "image_url"

    def test_normalize_headers_url_alias(self):
        mapping = _normalize_headers(["name", "url"])
        assert mapping.get("url") == "product_url"

    def test_normalize_headers_unrecognised_passthrough(self):
        mapping = _normalize_headers(["name", "random_column", "foo"])
        assert "random_column" not in mapping
        assert "foo" not in mapping

    def test_normalize_headers_case_insensitive(self):
        mapping = _normalize_headers(["NAME", "TITLE", "SKU"])
        assert mapping.get("NAME") == "name"
        assert mapping.get("TITLE") == "name"
        assert mapping.get("SKU") == "product_id"

    # ── CSV parser with aliases ──────────────────────────────────────────

    def test_csv_title_alias(self):
        csv_content = "title,price\nWidget,29.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert len(errors) == 0

    def test_csv_sku_alias(self):
        csv_content = "name,sku,price\nWidget,W-001,29.99"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["product_id"] == "W-001"
        assert len(errors) == 0

    def test_csv_availability_alias(self):
        csv_content = "name,price,availability\nWidget,29.99,in_stock"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["stock_status"] == "in_stock"
        assert len(errors) == 0

    def test_csv_image_alias(self):
        csv_content = "name,price,image\nWidget,29.99,https://example.com/img.jpg"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["image_url"] == "https://example.com/img.jpg"
        assert len(errors) == 0

    def test_csv_url_alias(self):
        csv_content = "name,price,url\nWidget,29.99,https://example.com/p"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["product_url"] == "https://example.com/p"
        assert len(errors) == 0

    def test_csv_all_aliases(self):
        csv_content = "title,sku,price,category,availability,desc,image,link\nWidget,W-001,29.99,Gadgets,in_stock,A widget,https://img.com,https://ex.com"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        p = products[0]
        assert p["name"] == "Widget"
        assert p["product_id"] == "W-001"
        assert p["price"] == 2999
        assert p["category"] == "Gadgets"
        assert p["stock_status"] == "in_stock"
        assert p["description"] == "A widget"
        assert p["image_url"] == "https://img.com"
        assert p["product_url"] == "https://ex.com"
        assert len(errors) == 0

    def test_csv_mixed_canonical_and_alias(self):
        """Mix canonical and alias headers in one file."""
        csv_content = "name,sku,price,availability\nWidget,W-001,29.99,in_stock"
        products, errors = parse_csv(csv_content)
        assert len(products) == 1
        assert products[0]["product_id"] == "W-001"
        assert products[0]["stock_status"] == "in_stock"
        assert len(errors) == 0

    # ── JSON parser with aliases ─────────────────────────────────────────

    def test_json_title_alias(self):
        data = '[{"title":"Widget","price":29.99}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert len(errors) == 0

    def test_json_sku_alias(self):
        data = '[{"name":"Widget","sku":"W-001"}]'
        products, errors = parse_json(data)
        assert len(products) == 1
        assert products[0]["product_id"] == "W-001"
        assert len(errors) == 0

    # ── XLSX parser with aliases ─────────────────────────────────────────

    def test_xlsx_title_alias(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title", "price"])
        ws.append(["Widget", "29.99"])
        path = tmp_path / "alias.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert len(errors) == 0

    def test_xlsx_all_aliases(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Title", "sku", "price"])
        ws.append(["Widget", "W-001", "29.99"])
        path = tmp_path / "alias2.xlsx"
        wb.save(str(path))

        products, errors = parse_xlsx(path)
        assert len(products) == 1
        assert products[0]["name"] == "Widget"
        assert products[0]["product_id"] == "W-001"
        assert len(errors) == 0


class TestImportCatalogEdgeCases:
    @pytest.fixture
    def tenant_id(self) -> UUID:
        return uuid4()

    def test_import_catalog_with_special_chars_in_values(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,product_id\nWidget<>&\"',W-001")

        db = MagicMock()
        with patch("app.knowledge.catalog.rag.index_products", return_value=1):
            imported, errors, batch_id = import_catalog(tenant_id, p, db)

        assert imported == 1
        assert db.add.call_count == 1

    def test_import_catalog_with_empty_batch_string_falls_back_to_timestamp(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\nWidget,29.99")

        db = MagicMock()
        with patch("app.knowledge.catalog.rag.index_products", return_value=1):
            imported, errors, batch_id = import_catalog(tenant_id, p, db, batch="")

        assert imported == 1
        assert batch_id.startswith("import_")  # empty batch falls back to timestamp

    def test_import_catalog_db_commit_error(self, tenant_id, tmp_path):
        p = tmp_path / "products.csv"
        p.write_text("name,price\nWidget,29.99")

        db = MagicMock()
        db.commit.side_effect = Exception("DB connection lost")

        with patch("app.knowledge.catalog.rag.index_products", return_value=1):
            with pytest.raises(Exception, match="DB connection lost"):
                import_catalog(tenant_id, p, db)
